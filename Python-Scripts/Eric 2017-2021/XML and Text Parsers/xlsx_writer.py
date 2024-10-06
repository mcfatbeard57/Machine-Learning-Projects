from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class XLSXWriter:
    def __init__(self, output_report):
        self.counter = 0
        self.blank_row = [' ']

    @staticmethod
    def font(size=9, bold=True, color='00000000'):
        return Font(name='Arial', size=size, bold=bold, italic=False, vertAlign='baseline', color=color)

    @staticmethod
    def alignment():
        return Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False, indent=0)

    @staticmethod
    def border(border_style='thin', color='FF000000'):
        return Border(left=Side(border_style=border_style, color=color),
                      right=Side(border_style=border_style, color=color),
                      top=Side(border_style=border_style, color=color),
                      bottom=Side(border_style=border_style, color=color),
                      diagonal=Side(border_style=None, color=color),
                      diagonal_direction=0,
                      outline=Side(border_style=border_style, color=color),
                      vertical=Side(border_style=border_style, color=color),
                      horizontal=Side(border_style=border_style, color=color))

    @staticmethod
    def pattern_fill(fill_type='solid', start_color='999999', end_color='999999'):
        return PatternFill(fill_type=fill_type, start_color=start_color, end_color=end_color)

    @staticmethod
    def format_header_cell(cell, font_size=9, font_bold=True, font_color='00FFFFFF'):
        cell.font = XLSXWriter.font(size=font_size, bold=font_bold, color=font_color)
        cell.border = XLSXWriter.border()
        cell.alignment = XLSXWriter.alignment()
        cell.fill = XLSXWriter.pattern_fill(start_color='00000000', end_color='00000000')

    @staticmethod
    def format_value_cell(c1, font_bold=True):
        c1.font = XLSXWriter.font(size=10, bold=font_bold)
        c1.alignment = XLSXWriter.alignment()

    @staticmethod
    def get_column_letter_from_number(col):
        result = ''
        while col > 0:
            col = col - 1
            remainder = col % 26
            digit = chr(int(remainder) + 97)
            result = f'{digit}{result}'
            col = (col - remainder) / 26
        return result

    def styled_header(self, sheet, header_data):
        for header in header_data:
            header = Cell(sheet, value=header)
            XLSXWriter.format_header_cell(header)
            yield header

        return header_data

    def styled_value(self, sheet, value_data):
        for value in value_data:
            value = Cell(sheet, value=value)
            XLSXWriter.format_value_cell(value, False)
            yield value

        return value_data

    def autofit_column_width(self, sheet):
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        for column_cells in sheet.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length

    def create_workbook(self):
        wb = Workbook()
        return wb

    def create_context_table(self, wb, arr_context):
        for context in arr_context:
            sheet = wb.create_sheet(context['sheet'], self.counter)
            sheet.append(self.styled_header(sheet, context['label']))
            sheet.append(self.blank_row)
            sheet.append(self.blank_row)
            sheet.append(self.styled_header(sheet, context['table_name']))
            sheet.append(self.styled_header(sheet, context['headers']))
            sheet.append(self.styled_value(sheet, context['values']))

            self.counter = self.counter + 1
            self.autofit_column_width(sheet)

    def create_nop_table(self, wb, dict_nop):
        sheet = wb.create_sheet(dict_nop['sheet'], self.counter)
        sheet.append(self.styled_header(sheet, dict_nop['label']))
        sheet.append(self.blank_row)
        nop_tables = dict_nop['nop_tables']
        for table in nop_tables:
            for key, value in table.items():
                if key == 'table_name':
                    sheet.append(self.styled_header(sheet, value))
                elif key == 'headers':
                    sheet.append(self.styled_header(sheet, value))
                elif key == 'values':
                    sheet.append(self.styled_value(sheet, value))

            sheet.append(self.blank_row)
        self.autofit_column_width(sheet)
        self.counter = self.counter + 1

    def create_generic_table(self, wb, dict_info):
        sheet = wb.create_sheet(dict_info['sheet'], self.counter)
        sheet.append(self.styled_header(sheet, dict_info['label']))
        sheet.append(self.blank_row)
        sheet_tables = dict_info['sheet_tables']
        for table in sheet_tables:
            for key, value in table.items():
                if key == 'table_name':
                    if value:
                        sheet.append(self.styled_header(sheet, value))
                elif key == 'headers':
                    sheet.append(self.styled_header(sheet, value))
                elif key == 'values':
                    for v in value:
                        if isinstance(v, list):
                            sheet.append(self.styled_value(sheet, v))

            sheet.append(self.blank_row)
        self.autofit_column_width(sheet)
        self.counter = self.counter + 1
