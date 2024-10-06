    
""" script to parse MIO input xml file every hour and append the output to the daily output file"""

import os
import xml.etree.ElementTree as ET
import openpyxl
import xlsxwriter
import operator 

import re
import sys
from datetime import datetime
import datetime
import shutil
import time
import logging


#####  hahhahahahaah !/opt/ericsson/rop_sftp/pyenv/bin/python hahaahahhahah
## =============================================================================================================
# Configurable variables
data_folder = '/opt/global/perf/opco1/mms/aggregated/moved'  # C:/Users/ezgupha/Desktop/PM XML parser
output_folder = '/opt/global/perf/opco1/mms/PM_Reports'

houseKeepingTime = 7 * 24 * 60 * 60

## =============================================================================================================

# Get latest file in the data folder fucntion
def find_latest_file(data_folder):
    files = [os.path.join(data_folder, x) for x in os.listdir(data_folder)]
    newest_dir = max(files , key = os.path.getctime)
    newest_filename = newest_dir.split('/')[-1] # put / insted of \\

    files.remove(newest_dir)

    newest2_dir = max(files , key = os.path.getctime)
    newest2_filename = newest2_dir.split('/')[-1] # put / insted of \\
    return newest_filename, newest_dir, newest2_filename, newest2_dir

## =============================================================================================================
# calling latest file AND SECOND LATEST FILE funtion
newest_filename, newest_dir, newest2_filename, newest2_dir = find_latest_file(data_folder)
output_filename = newest_filename.split('-')[0].split('.')[0]

# checking if its the first file of the day
is_first_file_of_the_day = False
if newest_filename.split('-')[0].split('.')[1] == '0000':
    is_first_file_of_the_day = True

# log_filename = newest_filename[:-6]

## =============================================================================================================
# creating logs
#MIO_PM_Log_DDMMYYYY
logFilename = ('MIO_PM_Log_{}.log'.format(newest_filename[:-6]))
logging.basicConfig(filename=logFilename, 
                    format='%(asctime)s %(message)s', 
                    filemode='w')
logger=logging.getLogger() 
logger.setLevel(logging.INFO) 


## =============================================================================================================
# Readnig xml file into ET
# Read file
def read_finput_file(filename):
    tree = ET.parse(filename)
    root = tree.getroot()
    return root

try:
    root = read_finput_file(newest_dir)
    logger.info('File {} read succesful '.format(newest_dir))
except Exception as e:
    logger.info('During reading new input file {} \n Found an error as \n  {}'.format(newest_dir, e))



## =============================================================================================================

# get index of each tag 
# MM1_receive_count = 0
# MM1_reject_count = 0
# MM4_receive_count = 0
# MM4_reject_count = 0
# MM3_receive_count = 0
# MM3_reject_count = 0
# MM7_receive_count = 0
# MM7_reject_count = 0
# MM3_incoming_messages = 0
# MM4_incoming_messages = 0
# MM7_incoming_messages = 0
try:
    for child in root:
        if child.tag == 'md':
            for child2 in child:
                if child2.tag == 'mi':
                    count = -3
                    for child3 in child2:
                        count +=1
                        if child3.text == 'total-number-of-MM1-received-messages':
                            MM1_receive_count = count
                        if child3.text == 'total-number-of-MM1-rejected-messages':
                            MM1_reject_count = count
                        if child3.text == 'total-number-of-MM4-received-messages':
                            MM4_receive_count = count                        
                        if child3.text == 'total-number-of-MM4-rejected-messages':
                            MM4_reject_count = count        
                        if child3.text == 'total-number-of-MM3-received-messages':
                            MM3_receive_count = count
                        if child3.text == 'total-number-of-MM3-rejected-messages':
                            MM3_reject_count = count
                        if child3.text == 'total-number-of-MM7-received-messages':
                            MM7_receive_count = count
                        if child3.text == 'total-number-of-MM7-rejected-messages':
                            MM7_reject_count = count
                        if child3.text == 'number-of-MM3-incoming-messages':
                            MM3_incoming_messages = count                  
                        if child3.text == 'number-of-MM4-incoming-messages':
                            MM4_incoming_messages = count                      
                        if child3.text == 'number-of-MM7-incoming-messages':
                            MM7_incoming_messages = count
    logger.info('parsing of new file {} \n is succesful. counts are following:'.format(newest_dir))
    logger.info('\n {} {} {} {} {} {} {} {} {} {} {} \n'.format(MM1_receive_count, MM1_reject_count,
                                                                                            MM4_receive_count, MM4_reject_count,
                                                                                            MM3_receive_count, MM3_reject_count,
                                                                                            MM7_receive_count, MM7_reject_count,
                                                                                            MM3_incoming_messages, MM4_incoming_messages,
                                                                                            MM7_incoming_messages))

except Exception as e:
    logger.info('During parsing of new file {} \n Found an error as \n  {}'.format(newest_dir, e))

# MM1_receive_count2 = 0
# MM1_reject_count2 = 0
# MM4_receive_count2 = 0
# MM4_reject_count2 = 0
# MM3_receive_count2 = 0
# MM3_reject_count2 = 0
# MM7_receive_count2 = 0
# MM7_reject_count2 = 0
# MM3_incoming_messages2 = 0
# MM4_incoming_messages2 = 0
# MM7_incoming_messages2 = 0

try:
    root2 = read_finput_file(newest2_dir)
    logger.info('File {} read succesful '.format(newest2_dir))
except Exception as e:
    logger.info('During reading old input file {} \n Found an error as \n  {}'.format(newest2_dir, e))


try:
    for child in root2:
        if child.tag == 'md':
            for child2 in child:
                if child2.tag == 'mi':
                    count = -3
                    for child3 in child2:
                        count +=1
                        if child3.text == 'total-number-of-MM1-received-messages':
                                        #  total-number-of-MM1-received-messages
                            MM1_receive_count2 = count
                        if child3.text == 'total-number-of-MM1-rejected-messages':
                            MM1_reject_count2 = count
                        if child3.text == 'total-number-of-MM4-received-messages':
                            MM4_receive_count2 = count                        
                        if child3.text == 'total-number-of-MM4-rejected-messages':
                            MM4_reject_count2 = count        
                        if child3.text == 'total-number-of-MM3-received-messages':
                            MM3_receive_count2 = count
                        if child3.text == 'total-number-of-MM3-rejected-messages':
                            MM3_reject_count2 = count
                        if child3.text == 'total-number-of-MM7-received-messages':
                            MM7_receive_count2 = count
                        if child3.text == 'total-number-of-MM7-rejected-messages':
                            MM7_reject_count2 = count
                        if child3.text == 'number-of-MM3-incoming-messages':
                            MM3_incoming_messages2 = count                  
                        if child3.text == 'number-of-MM4-incoming-messages':
                            MM4_incoming_messages2 = count                      
                        if child3.text == 'number-of-MM7-incoming-messages':
                            MM7_incoming_messages2 = count
    logger.info('parsing of old file {} \n is succesful. counts are following'.format(newest2_dir))
    logger.info('\n {} {} {} {} {} {} {} {} {} {} {} \n'.format(MM1_receive_count2, MM1_reject_count2,
                                                                                            MM4_receive_count2, MM4_reject_count2,
                                                                                            MM3_receive_count2, MM3_reject_count2,
                                                                                            MM7_receive_count2, MM7_reject_count2,
                                                                                            MM3_incoming_messages2, MM4_incoming_messages2,
                                                                                            MM7_incoming_messages2))
except Exception as e:
    logger.info('During parsing of file {} \n Found an error as \n  {}'.format(newest2_dir, e))
                    
## =============================================================================================================

# Get values and column names for excel file
def string_to_text(value1, value2):
    return int(value1) + int(value2)

def get_value_from_count(count,root):
    row_name  = root[1][-2][count+2].text
    value = string_to_text(root[1][-2][-1][count + 1].text, 
                            root[1][-2][-2][count + 1].text)
    return row_name, value

## =============================================================================================================
# gettting names and values of each required tag
try:
    # total-number-of-MM1-received-messages
    row_name1, value1 = get_value_from_count(MM1_receive_count,root)
    # total-number-of-MM1-rejected-messages
    row_name2, value2 = get_value_from_count(MM1_reject_count,root)
    # total-number-of-MM4-received-messages
    row_name3, value3 = get_value_from_count(MM4_receive_count,root)
    # total-number-of-MM4-rejected-messages
    row_name4, value4 = get_value_from_count(MM4_reject_count,root)
    # total-number-of-MM3-received-messages
    row_name5, value5 = get_value_from_count(MM3_receive_count,root)
    # total-number-of-MM3-rejected-messages
    row_name6, value6 = get_value_from_count(MM3_reject_count,root)
    # total-number-of-MM7-received-messages
    row_name7, value7 = get_value_from_count(MM7_receive_count,root)
    # total-number-of-MM7-rejected-messages
    row_name8, value8 = get_value_from_count(MM7_reject_count,root)
    # number-of-MM3-incoming-messages
    row_name9, value9 = get_value_from_count(MM3_incoming_messages,root)
    # number-of-MM4-incoming-messages
    row_name10, value10 = get_value_from_count(MM4_incoming_messages,root)
    # number-of-MM7-incoming-messages
    row_name11, value11 = get_value_from_count(MM7_incoming_messages,root)  


except Exception as e:
    logger.info('During parsing and getting values from index of file {} \n Found an error as \n  {}'.format(newest_dir, e))

try:
    # total-number-of-MM1-received-messages
    row_name1, value1_2 = get_value_from_count(MM1_receive_count2,root2)
    # total-number-of-MM1-rejected-messages
    row_name2, value2_2 = get_value_from_count(MM1_reject_count2,root2)
    # total-number-of-MM4-received-messages
    row_name3, value3_2 = get_value_from_count(MM4_receive_count2,root2)
    # total-number-of-MM4-rejected-messages
    row_name4, value4_2 = get_value_from_count(MM4_reject_count2,root2)
    # total-number-of-MM3-received-messages
    row_name5, value5_2 = get_value_from_count(MM3_receive_count2,root2)
    # total-number-of-MM3-rejected-messages
    row_name6, value6_2 = get_value_from_count(MM3_reject_count2,root2)
    # total-number-of-MM7-received-messages
    row_name7, value7_2 = get_value_from_count(MM7_receive_count2,root2)
    # total-number-of-MM7-rejected-messages
    row_name8, value8_2 = get_value_from_count(MM7_reject_count2,root2)
    # number-of-MM3-incoming-messages
    row_name9, value9_2 = get_value_from_count(MM3_incoming_messages2,root2)
    # number-of-MM4-incoming-messages
    row_name10, value10_2 = get_value_from_count(MM4_incoming_messages2,root2)
    # number-of-MM7-incoming-messages
    row_name11, value11_2 = get_value_from_count(MM7_incoming_messages2,root2)  

except Exception as e:
    logger.info('During parsing and getting values from index of file {} \n Found an error as \n  {}'.format(newest2_dir, e))


## =============================================================================================================

# Storing them into list for easy use
row_names = [row_name1,row_name2,row_name3,row_name4,row_name5,row_name6,row_name7,row_name8,row_name9,row_name10,row_name11]
value_list = [value1,value2,value3,value4,value5,value6,value7,value8,value9,value10,value11]
value_list2 = [value1_2,value2_2,value3_2,value4_2,value5_2,value6_2,value7_2,value8_2,value9_2,value10_2,value11_2]

logger.info('Current file row names are {} \n '.format(row_names))


column_name = ['0000','0100','0200','0300','0400','0500','0600','0700','0800',
            '0900','1000','1100','1200','1300','1400','1500','1600',
            '1700','1800','1900','2000','2100','2200','2300']

col_indx = list(range(1, 25))
row_index = list(range(1, 12))

## =============================================================================================================

try: 
    if is_first_file_of_the_day:
        final_value_list = value_list
    if not is_first_file_of_the_day:
        final_value_list = list(map(operator.sub, value_list, value_list2))
    logger.info('Computing the actual value from the cummulative values of file {} \n is a success \n orginal_value is {} \n past value is {} \n final value is {} \n '.format(newest_dir,value_list,value_list2,final_value_list))
except Exception as e:
    logger.info('During computing the actual value from the cummulative values of file {} \n Found an error as \n  {}'.format(newest_dir, e))


## =============================================================================================================

# Create new workbook at every 0000 hour
def creat_new_workbook():
    workbook = xlsxwriter.Workbook('MIO_PM_Reports_{}_output.xlsx'.format(output_filename)) 
    worksheet = workbook.add_worksheet()
    for i, col_name in zip(col_indx, column_name):
        worksheet.write(0, i, col_name)
    for i,row_nam in zip(row_index,row_names):
        worksheet.write(i, 0, row_nam)
    return worksheet,workbook

# populate data into workbook
def populate_data_into_workbook(worksheet):
    for i in range(len(column_name)):
        if column_name[i] == newest_filename.split('-')[0].split('.')[1]:
            for j,k in zip(range(1,12),final_value_list):
                worksheet.cell(row = j+1, column = i+2).value = k

## =============================================================================
# try creating the workbook then opening the existing workbook and populate it with parsed data, then save the file.
try:	
    if not os.path.isfile('MIO_PM_Reports_{}_output.xlsx'.format(output_filename)): 
        worksheet,workbook = creat_new_workbook()
        workbook.close()
    xfile = openpyxl.load_workbook('MIO_PM_Reports_{}_output.xlsx'.format(output_filename))
    sheet = xfile.get_sheet_by_name('Sheet1')
    populate_data_into_workbook(sheet)
    xfile.save('MIO_PM_Reports_{}_output.xlsx'.format(output_filename))
    logger.info('File MIO_PM_Reports_{}_output.xlsx write succesful '.format(output_filename))
except Exception as e:
    logger.info('During writing output file MIO_PM_Reports_{}_output.xlsx \n Found an error as \n  {}'.format(output_filename, e))

## =============================================================================

# perform housekeeping
# def houseKeeping(path):
#     now = time.time()
#     logger.info('Housekeeping begins----- \n ')
#     logger.info('Deleted files are \n ')
#     for file in os.listdir(path):
#         if not file.endswith(".py"):
#             if (file.endswith(".log") or file.endswith(".xlsx")):
#                 f = os.path.join(path, file)
#                 if os.stat(f).st_mtime < now - houseKeepingTime:
#                     logger.info(f)
#                     logger.info('\n')
#                     shutil.rmtree(f)
#     logger.info('Housekeeping finished \n')

# currently commented out
# houseKeeping(output_folder)

## =============================================================================

logger.info('Script ran successfully \n')

## =============================================================================